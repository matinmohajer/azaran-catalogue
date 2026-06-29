import json, os
from flask import Flask, render_template, jsonify
import openpyxl

app = Flask(__name__)
EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'محصولات_آذران_نور.xlsx')
VARIANT_ATTRS = ['color', 'size', 'light_color', 'power']

def load_groups():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws_p = wb['محصولات']
    ws_v = wb['واریانت‌ها']
    products = {}
    for row in ws_p.iter_rows(min_row=2, values_only=True):
        pid = row[0]
        if not pid:
            continue
        pfx = str(row[14]).strip() if row[14] else ''
        pg  = row[15]
        products[pid] = {
            'gid': pid,
            'category':   str(row[1])  if row[1]  else '',
            'name':       str(row[2])  if row[2]  else '',
            'code':       str(row[3])  if row[3]  else '',
            'model':      str(row[4])  if row[4]  else '',
            'dimensions': str(row[5])  if row[5]  else '',
            'lumen':      str(row[6])  if row[6]  else '',
            'material':   str(row[7])  if row[7]  else '',
            'cri':        str(row[8])  if row[8]  else '',
            'features':   str(row[9])  if row[9]  else '',
            'warranty':   str(row[10]) if row[10] else '',
            'date':       str(row[11]) if row[11] else '',
            'catalog':    str(row[12]) if row[12] else '',
            'notes':      str(row[13]) if row[13] else '',
            'image_url':  f'/static/images/pages/{pfx}_p{int(pg):02d}.jpg' if pfx and pg else '',
            'variants': [],
            'attr_options': {a: [] for a in VARIANT_ATTRS},
        }
    for row in ws_v.iter_rows(min_row=2, values_only=True):
        pid = row[0]
        if pid not in products:
            continue
        price = row[6]
        v = {
            'color':       str(row[2]) if row[2] else '',
            'size':        str(row[3]) if row[3] else '',
            'light_color': str(row[4]) if row[4] else '',
            'power':       str(row[5]) if row[5] else '',
            'price': int(price) if price and isinstance(price, (int, float)) else None,
            'availability': str(row[7]) if row[7] else '',
        }
        p = products[pid]
        p['variants'].append(v)
        for a in VARIANT_ATTRS:
            val = v[a]
            if val and val not in p['attr_options'][a]:
                p['attr_options'][a].append(val)
    groups = []
    for p in products.values():
        prices = [v['price'] for v in p['variants'] if v['price']]
        p['price_min'] = min(prices) if prices else None
        p['price_max'] = max(prices) if prices else None
        groups.append(p)
    return groups

GROUPS = load_groups()
GROUPS_JSON = json.dumps(GROUPS, ensure_ascii=False)

@app.route('/')
def index():
    return render_template('index.html', groups_json=GROUPS_JSON)

@app.route('/api/reload')
def reload():
    global GROUPS, GROUPS_JSON
    GROUPS = load_groups()
    GROUPS_JSON = json.dumps(GROUPS, ensure_ascii=False)
    return jsonify({'products': len(GROUPS), 'status': 'ok'})

if __name__ == '__main__':
    print(f'Loaded {len(GROUPS)} product groups')
    app.run(debug=True, port=8080, host='0.0.0.0')
